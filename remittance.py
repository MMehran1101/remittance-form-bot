from openpyxl import load_workbook


class Remittance:

    def __init__(
        self,
        co_name,
        city_name,
        product_name,
        small_num,
        large_num,
        xlarge_num,
    ):
        self.co_name = co_name
        self.city_name = city_name
        self.product_name = product_name
        self.small_num = small_num
        self.large_num = large_num
        self.xlarge_num = xlarge_num

    def create_remittance(self):
        # loading template excle file
        wb = load_workbook("template.xlsx")
        ws = wb.active

        cells = {
            "company_city": "A8",
            "small_product": "B14",
            "large_product": "B16",
            "xlarge_product": "B18",
            "small_num": "D14",
            "large_num": "D16",
            "xlarge_num": "D18",
            "warehouse_address": "A22",
        }
        data = {key: ws[addr].value for key, addr in cells.items()}

        new_company_city = data["company_city"].format(
            co_name=self.co_name, city=self.city_name
        )
        new_small_p = data["small_product"].format(product=self.product_name)
        new_large_p = data["large_product"].format(product=self.product_name)
        new_xlarge_p = data["xlarge_product"].format(product=self.product_name)

        new_small_num = data["small_num"].format(small_count=self.small_num)
        new_large_num = data["large_num"].format(large_count=self.large_num)
        new_xlarge_num = data["xlarge_num"].format(xlarge_count=self.xlarge_num)

        # new_werehouse_address (impeliment adresses)

        ws[cells["company_city"]].value = new_company_city
        ws[cells["small_product"]].value = new_small_p
        ws[cells["large_product"]].value = new_large_p
        ws[cells["xlarge_product"]].value = new_xlarge_p
        ws[cells["small_num"]].value = new_small_num
        ws[cells["large_num"]].value = new_large_num
        ws[cells["xlarge_num"]].value = new_xlarge_num

        # ws[werehouse_address].value = NEW ADDRESS LOAD FROM JSON

        wb.save("template_changed.xlsx")


# *** Cell details on template.xlsx ***
# [A8] co_name, city
# [B14][B16][B18] product
# [D14][D16][D18] small_count, large_count, xlarge_count
# [A22] werehouse_address
